#!/usr/bin/env python3
"""Excel to PDF Converter - LibreOffice or Python fallback"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from datetime import datetime

USAGE_LOGGER = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "usage_logger.py"))

try:
    subprocess.run([sys.executable, USAGE_LOGGER, "excel-to-pdf"], capture_output=True, timeout=5)
except Exception:
    pass


def find_libreoffice():
    """Detect LibreOffice executable path."""
    system = platform.system()
    if system == "Windows":
        candidates = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for c in candidates:
            if os.path.exists(c):
                return c
    elif system == "Darwin":
        path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        if os.path.exists(path):
            return path
    # Linux or fallback: try PATH
    try:
        subprocess.run(["soffice", "--version"], capture_output=True, timeout=5, check=True)
        return "soffice"
    except (FileNotFoundError, subprocess.SubprocessError):
        return None


def convert_with_libreoffice(input_path, output_dir=None, sheets=None):
    """Convert Excel to PDF using LibreOffice headless.

    Args:
        sheets: None=all sheets, list of sheet names or 0-based indices to convert.
                When specified, extracts target sheets to a temp file first.
    """
    soffice = find_libreoffice()
    if not soffice:
        return {"error": "LibreOffice not found"}

    abs_input = str(Path(input_path).absolute())
    if output_dir is None:
        output_dir = str(Path(input_path).parent.absolute())

    # If specific sheets requested, extract them to a temp file
    temp_file = None
    if sheets is not None:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(abs_input)
            target_names = _resolve_sheet_names(wb.sheetnames, sheets)
            # Remove non-target sheets
            for name in wb.sheetnames:
                if name not in target_names:
                    del wb[name]
            temp_file = os.path.join(output_dir, f"_temp_{Path(input_path).stem}.xlsx")
            wb.save(temp_file)
            wb.close()
            abs_input = temp_file
        except Exception as e:
            return {"error": f"Sheet extraction failed: {e}"}

    cmd = [soffice, "--headless", "--convert-to", "pdf", "--outdir", output_dir, abs_input]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)

    # Clean up temp file
    if temp_file and os.path.exists(temp_file):
        pdf_from_temp = os.path.join(output_dir, Path(temp_file).stem + ".pdf")
        final_pdf = os.path.join(output_dir, Path(input_path).stem + ".pdf")
        if os.path.exists(pdf_from_temp):
            os.replace(pdf_from_temp, final_pdf)
        os.remove(temp_file)

    if result.returncode != 0:
        return {"error": result.stderr or "LibreOffice conversion failed"}

    output_path = os.path.join(output_dir, Path(input_path).stem + ".pdf")
    if os.path.exists(output_path):
        return {"status": "success", "method": "libreoffice", "output": output_path}
    return {"error": "Output PDF not created"}


def _resolve_sheet_names(all_names, sheets):
    """Resolve sheet specifiers (names or 0-based indices) to a set of sheet names."""
    target = set()
    for s in sheets:
        if isinstance(s, int):
            if 0 <= s < len(all_names):
                target.add(all_names[s])
        elif s in all_names:
            target.add(s)
    return target


def convert_with_python(input_path, output_path=None, sheets=None):
    """Convert Excel to PDF using openpyxl + reportlab (data table only).

    Args:
        sheets: None=all sheets, list of sheet names or 0-based indices to convert.
    """
    try:
        from openpyxl import load_workbook
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    except ImportError as e:
        return {"error": f"Missing dependency: {e}. Install: pip install openpyxl reportlab"}

    if output_path is None:
        output_path = str(Path(input_path).with_suffix(".pdf"))

    # Register Japanese font
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))
        font_name = "HeiseiMin-W3"
    except Exception:
        font_name = "Helvetica"

    wb = load_workbook(input_path, data_only=True)
    styles = getSampleStyleSheet()
    story = []

    target_names = _resolve_sheet_names(wb.sheetnames, sheets) if sheets else None

    for sheet_name in wb.sheetnames:
        if target_names and sheet_name not in target_names:
            continue
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(c) if c is not None else "" for c in row])

        if not data:
            continue

        # Sheet title
        story.append(Paragraph(sheet_name, styles["Heading2"]))
        story.append(Spacer(1, 6))

        # Build table
        table = Table(data)
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(table)
        story.append(Spacer(1, 20))

    wb.close()

    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4))
    doc.build(story)

    return {"status": "success", "method": "python", "output": output_path}


def _parse_sheets_arg(arg):
    """Parse --sheets argument: comma-separated names or 0-based indices."""
    result = []
    for part in arg.split(","):
        part = part.strip()
        try:
            result.append(int(part))
        except ValueError:
            result.append(part)
    return result


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Convert Excel to PDF.")
    parser.add_argument("input", help="Input Excel file (.xlsx/.xls)")
    parser.add_argument("output", nargs="?", default=None, help="Output PDF path (default: same name .pdf)")
    parser.add_argument("--sheets", default=None,
                        help="Comma-separated sheet names or 0-based indices (default: all sheets). "
                             "Examples: --sheets 0  or  --sheets 'Sheet1,Sheet3'  or  --sheets 0,2")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(json.dumps({"error": f"File not found: {args.input}"}))
        sys.exit(1)

    sheets = _parse_sheets_arg(args.sheets) if args.sheets else None

    # Try LibreOffice first
    lo = find_libreoffice()
    if lo:
        output_dir = str(Path(args.output).parent) if args.output else None
        result = convert_with_libreoffice(args.input, output_dir, sheets=sheets)
    else:
        result = convert_with_python(args.input, args.output, sheets=sheets)

    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
